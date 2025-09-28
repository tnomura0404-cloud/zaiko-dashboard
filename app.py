import streamlit as st
import pandas as pd
from pathlib import Path
from datetime import datetime
from dateutil.relativedelta import relativedelta
import io
from openpyxl.styles import Font, PatternFill

# --------------------------------------------------------------------------------
# Excel整形関数 (v7.9版を完全に再現)
# --------------------------------------------------------------------------------
def format_excel_sheet_original(ws, df, columns_to_format, money_columns=None):
    is_shortage_report = ws.title.startswith("不足在庫")
    is_long_term_report = ws.title == "長期在庫リスト"
    
    for col_idx, column_cells in enumerate(ws.columns, 1):
        column_letter = column_cells[0].column_letter
        if (is_shortage_report and column_letter in ['C', 'D', 'E', 'F']) or \
           (is_long_term_report and column_letter in ['C', 'D', 'E', 'F']):
            ws.column_dimensions[column_letter].width = 15
        else:
            max_length = 0
            for cell in column_cells:
                if cell.value is not None: max_length = max(max_length, len(str(cell.value)))
            header_text = ws.cell(row=1, column=col_idx).value
            if header_text: max_length = max(max_length, len(str(header_text)))
            ws.column_dimensions[column_letter].width = max_length + 3

    header = [c.value for c in ws[1]]
    red_font = Font(color="FF0000")
    for col_name in columns_to_format:
        if col_name in header:
            col_idx = header.index(col_name) + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    if col_name == "差し引き数量" and cell.value < 0:
                        cell.font = red_font
    
    if money_columns:
        for col_name in money_columns:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
    return ws

# --------------------------------------------------------------------------------
# 分析ロジック (v7.9版をベースに)
# --------------------------------------------------------------------------------
def find_column_name(df_columns, possible_names):
    return next((name for name in possible_names if name in df_columns), None)

def analyze_inventory_original(src_file, rule_file, history_file):
    # (この関数のロジックは、前回のコードと全く同じです)
    ws_key = pd.read_excel(rule_file, sheet_name="キー", header=None, dtype=str).fillna("")
    key_dict = {str(val).strip(): str(ws_key.iloc[0, col_idx]).strip() for col_idx in range(ws_key.shape[1]) for val in ws_key.iloc[1:, col_idx] if str(val).strip()}
    manual_quantities = {}
    if "リスト" in pd.ExcelFile(rule_file).sheet_names:
        df_list = pd.read_excel(rule_file, sheet_name="リスト", dtype=str)
        quantity_col_name = find_column_name(df_list.columns, ["基準数量（手動）", "数量"])
        if quantity_col_name: manual_quantities = df_list.set_index('商品名')[quantity_col_name].apply(pd.to_numeric, errors='coerce').dropna().astype('Int64').to_dict()
    df_history = pd.DataFrame()
    if history_file is not None:
        try:
            df_history_raw = pd.read_excel(history_file, sheet_name='Data', engine='xlrd')
            column_rename_map = {"订单发行日": "order_date", "注文発行日": "order_date", "订单数量": "order_quantity", "注文数量": "order_quantity", "商品名称": "product_name"}
            df_history = df_history_raw.rename(columns=lambda c: column_rename_map.get(c, c))
            df_history["order_date"] = pd.to_datetime(df_history["order_date"], errors='coerce')
            df_history["order_quantity"] = pd.to_numeric(df_history["order_quantity"], errors='coerce')
            df_history.dropna(subset=["product_name", "order_date", "order_quantity"], inplace=True)
        except Exception as e: st.warning(f"発注履歴ファイルの読み込み中にエラー: {e}")
    ws_src = pd.read_excel(src_file, header=10, dtype=str).fillna("")
    inventory_col = find_column_name(ws_src.columns, ['客户在库', '在库数量', '在庫数量'])
    price_col = find_column_name(ws_src.columns, ['贩卖单价'])
    if not inventory_col:
        st.error("在庫数量列が見つかりません。")
        return [None]*7
    cols_to_map = {'商品名称': '商品名称', inventory_col: 'INVENTORY_LEVEL', '最终出荷日': '最终出荷日', price_col: '贩卖单价'}
    ws_src.rename(columns={k: v for k, v in cols_to_map.items() if k and k in ws_src.columns}, inplace=True)
    current_inventory_map = {str(row["商品名称"]).strip(): int(pd.to_numeric(row["INVENTORY_LEVEL"], errors='coerce')) for _, row in ws_src.iterrows() if pd.notna(row["商品名称"]) and pd.notna(row["INVENTORY_LEVEL"])}
    price_map = {str(row["商品名称"]).strip(): pd.to_numeric(row["贩卖单价"], errors='coerce') for _, row in ws_src.iterrows() if '贩卖单价' in ws_src.columns and pd.notna(row["商品名称"])}
    consumption_dict = {}
    if not df_history.empty:
        df_agg = df_history.groupby('product_name').agg(total_incoming=('order_quantity', 'sum'), first_order_date=('order_date', 'min')).reset_index()
        today = datetime.now()
        for _, row in df_agg.iterrows():
            p_name, total_in, first_date = row['product_name'], row['total_incoming'], row['first_order_date']
            months = max(1, (today.year - first_date.year) * 12 + (today.month - first_date.month))
            current_stock = current_inventory_map.get(p_name, 0)
            total_consumption = total_in - current_stock
            if total_consumption > 0:
                monthly_con = round(total_consumption / months)
                if monthly_con > 0: consumption_dict[p_name] = int(monthly_con)
    brand_groups = {}
    low_stock_auto, low_stock_manual, long_term_stock = set(), set(), set()
    one_year_ago = datetime.now() - relativedelta(years=1)
    for _, row in ws_src.iterrows():
        p_name = str(row["商品名称"]).strip()
        if not p_name: continue
        brand = next((bname for key, bname in key_dict.items() if key in p_name), "OTHER")
        if brand not in brand_groups: brand_groups[brand] = []
        brand_groups[brand].append(row.to_dict())
        inv_qty = current_inventory_map.get(p_name, 0)
        auto_qty = consumption_dict.get(p_name)
        manual_qty = manual_quantities.get(p_name)
        if auto_qty and inv_qty < auto_qty: low_stock_auto.add(p_name)
        if manual_qty and inv_qty < manual_qty: low_stock_manual.add(p_name)
        ship_date = pd.to_datetime(str(row["最终出荷日"]).strip(), errors='coerce')
        if pd.notna(ship_date) and ship_date < one_year_ago: long_term_stock.add(p_name)
    report_items_auto, report_items_manual, long_term_items_full = [], [], []
    df_src_for_report = ws_src.drop_duplicates(subset=['商品名称'])
    for _, row in df_src_for_report.iterrows():
        p_name = str(row['商品名称']).strip()
        if not p_name: continue
        brand_name = next((bname for key, bname in key_dict.items() if key in p_name), "OTHER")
        inv_qty = current_inventory_map.get(p_name, 0)
        auto_qty = consumption_dict.get(p_name)
        manual_qty = manual_quantities.get(p_name)
        if p_name in low_stock_auto: report_items_auto.append({"ブランド": brand_name, "商品名": p_name, "在庫数": inv_qty, "基準数量(自動)": auto_qty, "基準数量(手動)": manual_qty, "差し引き数量": inv_qty - (auto_qty or 0)})
        if p_name in low_stock_manual: report_items_manual.append({"ブランド": brand_name, "商品名": p_name, "在庫数": inv_qty, "基準数量(自動)": auto_qty, "基準数量(手動)": manual_qty, "差し引き数量": inv_qty - (manual_qty or 0)})
        if p_name in long_term_stock:
            ship_date = pd.to_datetime(row['最终出荷日'], errors='coerce')
            price = price_map.get(p_name)
            total_amount = inv_qty * price if pd.notna(price) and inv_qty > 0 else None
            long_term_items_full.append({"ブランド": brand_name, "商品名": p_name, "最終出荷日": ship_date.date() if pd.notna(ship_date) else None, "経過日数": (datetime.now() - ship_date).days if pd.notna(ship_date) else 0, "在庫数": inv_qty, "贩卖单价": price, "合計金額": total_amount})
    return brand_groups, low_stock_auto, low_stock_manual, long_term_stock, pd.DataFrame(report_items_auto), pd.DataFrame(report_items_manual), pd.DataFrame(long_term_items_full)

# --- Excelダウンロード用の関数 (v7.9版を完全に再現) ---
def to_excel_original(brand_groups, low_auto, low_manual, long_term, df_auto, df_manual, df_long):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # 1. レポートシートの作成
        comma_cols = ["在庫数", "基準数量(自動)", "基準数量(手動)", "差し引き数量", "経過日数"]
        if not df_auto.empty:
            df_auto_sorted = df_auto.sort_values(by=["ブランド", "商品名"])
            df_auto_sorted.to_excel(writer, sheet_name="不足在庫(自動ベース)", index=False)
            format_excel_sheet_original(writer.sheets["不足在庫(自動ベース)"], df_auto_sorted, comma_cols)
        if not df_manual.empty:
            df_manual_sorted = df_manual.sort_values(by=["ブランド", "商品名"])
            df_manual_sorted.to_excel(writer, sheet_name="不足在庫(手動ベース)", index=False)
            format_excel_sheet_original(writer.sheets["不足在庫(手動ベース)"], df_manual_sorted, comma_cols)
        if not df_long.empty:
            df_long_sorted = df_long.sort_values(by=["経過日数"], ascending=False)
            df_long_sorted = df_long_sorted[["ブランド", "商品名", "最終出荷日", "経過日数", "在庫数", "贩卖单价", "合計金額"]]
            df_long_sorted.to_excel(writer, sheet_name="長期在庫リスト", index=False)
            ws = writer.sheets["長期在庫リスト"]
            format_excel_sheet_original(ws, df_long_sorted, ["在庫数", "経過日数"], money_columns=["贩卖单价", "合計金額"])
        
        # 2. ブランド別シートの作成とハイライト
        low_fill, long_fill = PatternFill(fill_type="solid", fgColor="FFFF00"), PatternFill(fill_type="solid", fgColor="FFCCCC")
        for brand in sorted(brand_groups.keys()):
            df_brand = pd.DataFrame(brand_groups[brand])
            
            # --- ★★★改善点1：不要な列を完全に削除★★★ ---
            cols_to_drop = ['购入单价', '贩卖单价.2', '金额(USD)', '贩卖单价.3', '金额(HKD)', '备注', 'brand', 'INVENTORY_LEVEL', '贩卖单价']
            df_brand_cleaned = df_brand.drop(columns=[col for col in cols_to_drop if col in df_brand.columns], errors='ignore')
            
            # --- ★★★改善点2：日付から時刻(00:00:00)を削除★★★ ---
            for col in ['受注月日', '最终出荷日']:
                if col in df_brand_cleaned.columns:
                    df_brand_cleaned[col] = pd.to_datetime(df_brand_cleaned[col], errors='coerce').dt.date
            
            df_brand_cleaned.to_excel(writer, sheet_name=brand, index=False)
            ws = writer.sheets[brand]
            format_excel_sheet_original(ws, df_brand_cleaned, [])
            
            header = [cell.value for cell in ws[1]]
            try:
                p_idx, s_idx = header.index("商品名称") + 1, header.index("最终出荷日") + 1
                for r_idx in range(2, ws.max_row + 1):
                    p_cell = ws.cell(row=r_idx, column=p_idx)
                    p_name = str(p_cell.value)
                    if p_name in low_auto or p_name in low_manual:
                        p_cell.fill = low_fill
                    if p_name in long_term:
                        ws.cell(row=r_idx, column=s_idx).fill = long_fill
            except (ValueError, AttributeError): pass

    return output.getvalue()

# --------------------------------------------------------------------------------
# Streamlit UI部分 (変更なし)
# --------------------------------------------------------------------------------
st.set_page_config(layout="wide")
st.title('📈 在庫分析ダッシュボード')
BASE_PATH = Path(__file__).resolve().parent
DEFAULT_RULE_FILE = BASE_PATH / "振り分けルール.xlsx"
DEFAULT_HISTORY_FILE = BASE_PATH / "発注履歴.xls"
if not DEFAULT_RULE_FILE.exists():
    st.error("エラー: アプリに「振り分けルール.xlsx」が同梱されていません。")
    st.stop()
rule_file = DEFAULT_RULE_FILE
history_file = DEFAULT_HISTORY_FILE if DEFAULT_HISTORY_FILE.exists() else None
st.info("👇 分析したい「元在庫表」をアップロードしてください。")
uploaded_src_file = st.file_uploader("", type=['xlsx', 'xls'], label_visibility="collapsed")
st.sidebar.header("⚙️ 設定")
st.sidebar.markdown("このダッシュボードは、同梱されたマスターファイルを使用します。")
with st.sidebar.expander("もし、特別なファイルで試したい場合はこちら"):
    uploaded_rule_override = st.file_uploader("特別な「振り分けルール」")
    uploaded_history_override = st.file_uploader("特別な「発注履歴」")
    if uploaded_rule_override: rule_file = uploaded_rule_override
    if uploaded_history_override: history_file = uploaded_history_override
if uploaded_src_file:
    st.success(f"「{uploaded_src_file.name}」を読み込みました。")
    with st.spinner('在庫データを分析中...'):
        brand_groups, low_auto, low_manual, long_term, df_auto, df_manual, df_long = analyze_inventory_original(uploaded_src_file, rule_file, history_file)
    if brand_groups is not None:
        st.success('分析が完了しました！')
        st.header('分析結果')
        excel_data = to_excel_original(brand_groups, low_auto, low_manual, long_term, df_auto, df_manual, df_long)
        st.download_button(label="📄 見やすいExcel形式で全データをダウンロード", data=excel_data, file_name=f"在庫レポート_{datetime.now().strftime('%Y%m%d')}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        tab1, tab2, tab3 = st.tabs([f"不足在庫(自動) ({len(df_auto)})", f"不足在庫(手動) ({len(df_manual)})", f"長期在庫 ({len(df_long)})"])
        with tab1: st.dataframe(df_auto)
        with tab2: st.dataframe(df_manual)
        with tab3: st.dataframe(df_long)
        st.divider()
        st.header('全在庫リスト（ブランド別詳細）')
        df_full = pd.concat([pd.DataFrame(v) for v in brand_groups.values()])
        brand_list = ["全ブランド表示"] + sorted(list(brand_groups.keys()))
        selected_brand = st.selectbox('表示したいブランドを選択してください:', brand_list)
        if selected_brand == "全ブランド表示":
            st.dataframe(df_full.drop(columns=['brand'], errors='ignore'))
        else:
            st.dataframe(pd.DataFrame(brand_groups[selected_brand]).drop(columns=['brand'], errors='ignore'))
